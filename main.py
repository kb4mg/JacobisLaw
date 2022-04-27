# This script to compute values for the Maximum Power Transfer Theorem known as Jabobi's Law
# Author: Martin Buehring - KB4MG
# April , 2022
# v1 -excel interface - branched code
# v2 - improved code to commit
# Program Dependencies:
# 1) requires openpyxl package to be installed

import openpyxl
from openpyxl import Workbook


# Program parameters to be set
sourceVoltage = 100  # in volts , eg 100 volts
sourceResistance = 50 # in Ohms
loadResistanceStartVal = 10  #in Ohms, but never zero or infinity
loadResistanceEndVal = 200  # in Ohms
loadResistanceStepVal = 5  #in Ohms, to add and sweep the range

# program variables
loadResistance = 0
loopCurrent = 0.0  # Current in the load loop
loadVoltage = 0.0  # Voltage across the Load
loadPower = 0.0  # Total power in the load
maxPower = 0.0  # Maximum Power seen

# Functions
def printBanner(pgname):
    # Use a breakpoint in the code line below to debug your script.
    print(f'{pgname}')  # Press Ctrl+F8 to toggle the breakpoint.

def powerVal(i,v):
    return (i * v)

# Excel Interface - create sheet so we can graph it.
def writeExcel():
    ws['A1'] = 'Load Resistance'
    ws['B1'] = 'Load Power'
    ws['C1'] = 'Loop Current'
    ws['D1'] = 'Load Voltage'
    rowStart = 2 # because row 1 is the header
    columnStart = 1
    #Fill in the Excel table based on how many datapoints we have in our lists
    for i in range(0, len(powerList)):
        ws.cell(rowStart + i, columnStart).value = loadResList[i]
        ws.cell(rowStart + i, columnStart + 1).value = powerList[i]
        ws.cell(rowStart + i, columnStart + 2).value = loopCurList[i]
        ws.cell(rowStart + i, columnStart + 3).value = loopVoltsList[i]

    wb.save('MPTT.xlsx')  # Save the file in Excel format

###############################################
## MAIN PROGRAM
###############################################
# Lists to keep results of calculations
loadResList = []
loopCurList = []
loopVoltsList = []
powerList = []


if __name__ == '__main__':
    printBanner('MPTT Calculation and creaton of Excel file')
    #Establish objects for Excel interface
    wb = Workbook()
    ws = wb['Sheet']

# Compute power for various load resistances - we are rounding all floats to three decimal places
    while loadResistance < loadResistanceEndVal:
        loadResistance = loadResistance + loadResistanceStepVal
        print('* At Load = ',loadResistance)
        loadResList.append(loadResistance)  #add it to the list of loads
        loopCurrent = sourceVoltage/(loadResistance + sourceResistance)  # add load plus source resistances
        loopCurList.append(round(loopCurrent,3))
        print(f' ','LoopCurrent:',end="")
        print('%.3f'%loopCurrent, 'A') # Format current to three decimal places
        loadVoltage = loopCurrent * loadResistance  # Formula V = I*R
        loopVoltsList.append(round(loadVoltage,3))  #add it to the list of voltages
        print(f' ','Load Voltage:',end="")
        print('%.3f'%loadVoltage, 'V')
        loadPower = powerVal(loopCurrent, loadVoltage)  # Formula P= I**2 * R
        powerList.append(round(loadPower,3))  #add to the power list
        print(f' ','Load Power:', end="")
        print('%.3f'%loadPower, 'W with Load Resistance:', loadResistance,'\n')
        #Report the highest power value from all calculations
        if (maxPower < loadPower):
            maxPower = loadPower
print('Computed Max Power is ', maxPower)

# Call the Write EXCEL Function
writeExcel()
print('Generation Complete ....')